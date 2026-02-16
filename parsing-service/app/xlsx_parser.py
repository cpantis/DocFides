"""XLSX/XLS spreadsheet parsing using openpyxl."""

from typing import Optional
from app.models import TableData, MergedCell


def parse_xlsx(content: bytes, filename: str) -> dict:
    """Parse Excel spreadsheets into structured table data."""
    import io
    import openpyxl

    tables: list[TableData] = []
    raw_texts: list[str] = []

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=False)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row is None or ws.max_column is None:
                continue
            if ws.max_row < 1:
                continue

            headers: list[str] = []
            rows: list[list[str]] = []
            merged: list[MergedCell] = []

            # Extract merged cells info
            for merge_range in ws.merged_cells.ranges:
                merged.append(MergedCell(
                    row=merge_range.min_row - 1,
                    col=merge_range.min_col - 1,
                    row_span=merge_range.max_row - merge_range.min_row + 1,
                    col_span=merge_range.max_col - merge_range.min_col + 1,
                ))

            # Read all rows
            all_rows: list[list[str]] = []
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
                all_rows.append([_cell_to_str(cell) for cell in row])

            if not all_rows:
                continue

            # First row = headers
            headers = all_rows[0]
            rows = all_rows[1:] if len(all_rows) > 1 else []

            # Build raw text representation
            raw_texts.append(f"--- Sheet: {sheet_name} ---")
            for row_data in all_rows:
                raw_texts.append("\t".join(row_data))

            tables.append(TableData(
                headers=headers,
                rows=rows,
                merged_cells=merged,
                confidence=98.0,  # Native parsing = high confidence
            ))

        wb.close()
    except Exception as e:
        return {
            "tables": [],
            "raw_text": f"[Excel parse error: {str(e)}]",
            "error": str(e),
        }

    return {
        "tables": tables,
        "raw_text": "\n".join(raw_texts),
        "page_count": len(tables),
    }


def _cell_to_str(cell: Optional[object]) -> str:
    """Convert cell value to string, handling None and numeric types."""
    if cell is None:
        return ""
    if isinstance(cell, float):
        # Preserve integer formatting when appropriate
        if cell == int(cell):
            return str(int(cell))
        return str(cell)
    return str(cell)
